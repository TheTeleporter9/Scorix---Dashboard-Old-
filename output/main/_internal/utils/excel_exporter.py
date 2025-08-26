import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def export_games_to_excel(games, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Games'

    # Header
    headers = [
        'Game Number', 'Timestamp',
        'Team 1 Name', 'Team 1 Score', 'Team 1 Orange Balls', 'Team 1 Purple Balls',
        'Team 2 Name', 'Team 2 Score', 'Team 2 Orange Balls', 'Team 2 Purple Balls'
    ]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    for game in games:
        game_number = game.get('GameNumber', '')
        timestamp = game.get('timestamp', '')
        t1 = game.get('Team1', {})
        t2 = game.get('Team2', {})
        row = [
            game_number,
            timestamp,
            t1.get('Name', ''), t1.get('Score', ''), t1.get('OrangeBalls', ''), t1.get('PurpleBalls', ''),
            t2.get('Name', ''), t2.get('Score', ''), t2.get('OrangeBalls', ''), t2.get('PurpleBalls', '')
        ]
        ws.append(row)
        # Highlight winner
        if t1.get('Score', 0) > t2.get('Score', 0):
            ws.cell(row=ws.max_row, column=3).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        elif t2.get('Score', 0) > t1.get('Score', 0):
            ws.cell(row=ws.max_row, column=7).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    wb.save(filename) 
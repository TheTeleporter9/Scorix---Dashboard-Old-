import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def export_games_to_excel(games, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Games'

    # Header
    headers = [
        'Game Number', 'Timestamp',
        'Team 1 Name', 'Team 1 Score', 'Team 1 Orange Balls', 'Team 1 Purple Balls',
        'Team 2 Name', 'Team 2 Score', 'Team 2 Orange Balls', 'Team 2 Purple Balls'
    ]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    # Add game data
    for game in games:
        row = []
        game_number = game.get('GameNumber', '')
        timestamp = game.get('timestamp', '')
        t1 = game.get('Team1', {})
        t2 = game.get('Team2', {})
        
        row.extend([
            game_number,
            timestamp,
            t1.get('Name', ''),
            t1.get('Score', ''),
            t1.get('Orange', ''),
            t1.get('Purple', ''),
            t2.get('Name', ''),
            t2.get('Score', ''),
            t2.get('Orange', ''),
            t2.get('Purple', '')
        ])
        
        ws.append(row)
        
    # Save workbook
    wb.save(filename)

def export_excel():
    """Export all games to Excel"""
    from data.mongodb_client import collection
    from tkinter import filedialog
    games = list(collection.find())
    if not games:
        return None
    
    filename = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx')]
    )
    if not filename:
        return None
        
    export_games_to_excel(games, filename)
    return filename

def export_schedule_csv(schedule):
    """Export the schedule to a CSV file"""
    from tkinter import filedialog
    filename = filedialog.asksaveasfilename(
        defaultextension='.csv',
        filetypes=[('CSV files', '*.csv')]
    )
    if not filename:
        return None
        
    with open(filename, 'w') as f:
        f.write('Match,Team 1,Team 2,Played\n')
        for i, m in enumerate(schedule['matches']):
            team1 = m['team1']
            team2 = m['team2']
            played = 'Yes' if m['played'] else 'No'
            f.write(f'{i+1},{team1},{team2},{played}\n')
            
    return filename

